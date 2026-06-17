"""Excel exporter — openpyxl based .xlsx with formatted tables."""

from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..core.model import UnifiedProject
from .base import BaseExporter


HEADER_FILL = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)

HEADERS = [
    "Track", "Track Type", "#", "Clip Name", "Source File",
    "Source In", "Source Out", "Record In", "Record Out",
    "Duration", "Speed %", "Transition", "Transition Duration",
]


class ExcelExporter(BaseExporter):
    @classmethod
    def format_name(cls) -> str:
        return "Microsoft Excel (xlsx)"

    @classmethod
    def file_extension(cls) -> str:
        return ".xlsx"

    def export(self, project: UnifiedProject, output_path: Path) -> Path:
        wb = Workbook()

        for idx, timeline in enumerate(project.timelines):
            ws_name = f"Timeline_{idx + 1}"[:31]
            ws = wb.create_sheet(title=ws_name) if idx > 0 else wb.active
            ws.title = ws_name

            # Metadata
            ws.cell(row=1, column=1, value=f"Timeline: {timeline.name}").font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f"Framerate: {float(timeline.framerate):.4f} fps")
            ws.cell(row=2, column=2, value=f"Drop Frame: {timeline.drop_frame}")
            ws.cell(row=3, column=1, value=f"Project: {project.metadata.title}")

            # Header row (row 5)
            header_row = 5
            for col, h in enumerate(HEADERS, 1):
                cell = ws.cell(row=header_row, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

            # Data rows
            row_num = header_row + 1
            for track in timeline.tracks:
                for clip in track.clips:
                    vals = [
                        track.name,
                        track.track_type.value,
                        row_num - header_row,
                        clip.clip_name,
                        clip.source_file or "",
                        clip.source_in.to_smpte_string() if clip.source_in else "",
                        clip.source_out.to_smpte_string() if clip.source_out else "",
                        clip.record_in.to_smpte_string() if clip.record_in else "",
                        clip.record_out.to_smpte_string() if clip.record_out else "",
                        clip.record_duration.to_smpte_string() if clip.record_duration else "",
                        clip.speed,
                        clip.transition.transition_type.value if clip.transition else "Cut",
                        clip.transition.duration.to_smpte_string() if clip.transition and clip.transition.duration else "",
                    ]
                    for col, v in enumerate(vals, 1):
                        cell = ws.cell(row=row_num, column=col, value=v)
                        cell.border = THIN_BORDER
                    row_num += 1

                # Add blank row between tracks
                row_num += 1

            # Column widths
            widths = [16, 12, 6, 22, 36, 12, 12, 12, 12, 12, 10, 12, 18]
            for col, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{col}"].width = w

        # Remove default sheet if multiple created
        if len(wb.sheetnames) > len(project.timelines):
            del wb["Sheet"]

        wb.save(str(output_path))
        return output_path


# Register
from .base import ExporterRegistry
ExporterRegistry.register(ExcelExporter)
