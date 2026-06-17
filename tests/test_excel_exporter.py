"""Tests for Excel exporter."""

from pathlib import Path
import tempfile
from deepworld.parsers import ParserRegistry
from deepworld.exporters import ExporterRegistry
from openpyxl import load_workbook

SAMPLES = Path(__file__).parent.parent / "samples"
SAMPLE_EDL = SAMPLES / "sample_cmx3600.edl"


class TestExcelExporter:
    def setup_method(self):
        self.parser = ParserRegistry.get_parser(SAMPLE_EDL)
        self.exporter = ExporterRegistry.get_exporter(Path("test.xlsx"))
        self.project = self.parser.parse(SAMPLE_EDL)

    def test_exporter_type(self):
        assert "Excel" in self.exporter.format_name()

    def test_export_creates_file(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        try:
            result = self.exporter.export(self.project, path)
            assert result.exists()
            assert result.stat().st_size > 0
        finally:
            path.unlink(missing_ok=True)

    def test_export_has_sheets(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        try:
            self.exporter.export(self.project, path)
            wb = load_workbook(path)
            assert len(wb.sheetnames) >= 1
            wb.close()
        finally:
            path.unlink(missing_ok=True)

    def test_export_has_data(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        try:
            self.exporter.export(self.project, path)
            wb = load_workbook(path)
            ws = wb.active
            assert ws.max_row > 5  # Has header + data + metadata rows
            assert ws.max_column >= 10
            wb.close()
        finally:
            path.unlink(missing_ok=True)
